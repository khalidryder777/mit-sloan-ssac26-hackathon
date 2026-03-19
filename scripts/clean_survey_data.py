"""
Clean the Wasserman x Sloan MIT survey cross-tabulation data.

Input:  Power Players Research - Wasserman x Sloan MIT research.xlsx
Output: data/cleaned/survey_crosstabs_long.csv   (long-format, analysis-ready)
        data/cleaned/survey_crosstabs_wide.csv   (wide-format, one row per question-response)
        data/cleaned/survey_freetext.csv         (open-ended verbatim responses, separated out)
"""

import re
import pandas as pd
import openpyxl
from pathlib import Path

# ---------- paths ----------
ROOT = Path(__file__).resolve().parent.parent
INPUT = ROOT / "problem and data" / "Power Players Research - Wasserman x Sloan MIT research.xlsx"
OUT_DIR = ROOT / "data" / "cleaned"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ---------- column mapping ----------
# Row 2-3 headers define the cross-tab segments:
#   D(4)-F(6):  Moms -> Total, Mom, Non-Moms
#   H(8)-I(9):  Women Sports Fans -> Total, Women Sports Fans
#   K(11)-R(18): Age -> Total, Under 18, 18-28, 29-44, 35-44, 45-60, 61-65, 66+

SEGMENT_COLS = {
    # (col_index_0based, segment_group, segment_value)
    (3, "moms", "total"):           "moms_total",
    (4, "moms", "mom"):             "moms_mom",
    (5, "moms", "non_mom"):         "moms_non_mom",
    (7, "women_sports_fans", "total"):     "wsf_total",
    (8, "women_sports_fans", "women_sports_fans"): "wsf_women_sports_fans",
    (10, "age", "total"):           "age_total",
    (11, "age", "under_18"):        "age_under_18",
    (12, "age", "18_28"):           "age_18_28",
    (13, "age", "29_44"):           "age_29_44",  # note: age header says 29-44
    (14, "age", "35_44"):           "age_35_44",
    (15, "age", "45_60"):           "age_45_60",
    (16, "age", "61_65"):           "age_61_65",
    (17, "age", "66_plus"):         "age_66_plus",
}

WIDE_COLS = {idx: label for (idx, _, _), label in SEGMENT_COLS.items()}


def extract_question_id(text):
    """Extract question ID like 'Q1', 'Q8_0_1_RANK' from question text."""
    if not text:
        return None
    m = re.match(r"(Q\d+(?:_\d+_\d+_RANK|_\d+_TEXT)?)", text)
    return m.group(1) if m else None


def extract_question_text(text):
    """Extract the clean question text from the full label."""
    if not text:
        return None
    # Remove the leading "Q1: " or "Q8_0_1_RANK:\n\nQ: " prefix
    # Pattern 1: "Q1: question text"
    m = re.match(r"Q\d+(?:_\d+_\d+_RANK|_\d+_TEXT)?[:\s]*(?:\n*Q:\s*)?(.+)", text, re.DOTALL)
    if m:
        cleaned = m.group(1).strip()
        # For RANK questions, the ANSWER part is embedded - extract question only
        if "ANSWER:" in cleaned:
            cleaned = cleaned[:cleaned.index("ANSWER:")].strip().rstrip("\n")
        # Collapse whitespace
        cleaned = re.sub(r"\s+", " ", cleaned)
        return cleaned
    return text.strip()


def extract_rank_answer(text):
    """For RANK questions, extract the answer option from the question label."""
    if not text or "ANSWER:" not in text:
        return None
    m = re.search(r"ANSWER:\s*(.+?)(?:\s*-\s*Rank\s*)?$", text, re.DOTALL)
    if m:
        return re.sub(r"\s+", " ", m.group(1).strip().rstrip("- Rank").strip())
    return None


def load_raw_rows(ws):
    """Load all rows from the worksheet into a list of dicts."""
    rows = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        rows.append({
            "col_a": row[0],   # question label
            "col_b": row[1],   # response option / sub-item
            "col_c": row[2],   # rank or scale (for matrix questions)
            "values": [row[i] if i < len(row) else None for i in range(3, 18)],
        })
    return rows


def is_blank_row(r):
    """Check if a row is a separator (all None or empty string)."""
    if r["col_a"] is not None and str(r["col_a"]).strip():
        return False
    if r["col_b"] is not None and str(r["col_b"]).strip():
        return False
    if r["col_c"] is not None and str(r["col_c"]).strip():
        return False
    return True


def is_skip_row(r):
    """Rows to exclude: blank separators, N/A padding, Top/Bottom Box summaries, 'No data' messages."""
    if is_blank_row(r):
        return True
    b = str(r["col_b"]).strip() if r["col_b"] else ""
    if b == "N/A":
        return True
    if b in ("Top Box", "Bottom Box"):
        return True
    a = str(r["col_a"]).strip() if r["col_a"] else ""
    if "No data to export" in a or "No data to export" in b:
        return True
    return False


def parse_all_questions(raw_rows):
    """
    Parse the raw rows into a flat list of records.
    Each record: {question_id, question_text, response_option, rank_or_scale, sub_item, ...values}
    """
    records = []
    current_question_label = None
    current_question_id = None
    current_question_text = None
    current_rank_answer = None  # for RANK-style questions where answer is in col A label
    current_sub_item = None     # for matrix questions (col B holds sub-item, col C holds scale)

    for r in raw_rows:
        if is_skip_row(r):
            # Reset sub-item on blank rows (matrix question sub-items are separated by blanks)
            if is_blank_row(r):
                current_sub_item = None
            continue

        # If col A has a value, it's a new question block
        if r["col_a"] is not None and str(r["col_a"]).strip():
            current_question_label = str(r["col_a"]).strip()
            current_question_id = extract_question_id(current_question_label)
            current_question_text = extract_question_text(current_question_label)
            current_rank_answer = extract_rank_answer(current_question_label)
            current_sub_item = None

        # Determine the response option and any sub-item/scale
        col_b = r["col_b"]
        col_c = r["col_c"]

        # Convert to string, handle None
        col_b_str = str(col_b).strip() if col_b is not None else None
        col_c_str = str(col_c).strip() if col_c is not None else None

        # Skip empty col_b with no col_c
        if not col_b_str and not col_c_str:
            continue

        # Determine the structure type:
        # Type A: RANK questions (Q8_*, Q11_*, Q24_*, Q25_*) - answer in question label, rank in col_b
        # Type B: Matrix (Q21, Q39) - sub-item in col_b, scale/rank in col_c
        # Type C: Simple - response in col_b, col_c empty

        if current_rank_answer:
            # Type A: RANK question
            response_option = col_b_str  # "Rank Number 1", "Rank Number 2", etc.
            sub_item = current_rank_answer
            rank_or_scale = None
        elif col_c_str and col_b_str:
            # Type B: Matrix - both col B and C have values, new sub-item
            current_sub_item = col_b_str
            sub_item = current_sub_item
            response_option = col_c_str
            rank_or_scale = col_c_str
        elif col_c_str and not col_b_str:
            # Type B continued: sub-item carries forward from previous row
            sub_item = current_sub_item
            response_option = col_c_str
            rank_or_scale = col_c_str
        else:
            # Type C: Simple
            response_option = col_b_str
            sub_item = None
            rank_or_scale = None

        # Build value dict
        val_dict = {}
        for i, v in enumerate(r["values"]):
            col_idx = i + 3  # values start at col index 3
            if col_idx in WIDE_COLS:
                val_dict[WIDE_COLS[col_idx]] = v if v is not None else 0

        records.append({
            "question_id": current_question_id,
            "question_text": current_question_text,
            "sub_item": sub_item,
            "response_option": response_option,
            **val_dict,
        })

    return records


def build_wide_df(records):
    """Build the wide-format DataFrame."""
    df = pd.DataFrame(records)

    # Fill numeric columns: convert to int where possible
    num_cols = [c for c in df.columns if c.startswith(("moms_", "wsf_", "age_"))]
    for c in num_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype(int)

    # Clean text columns
    for c in ["question_id", "question_text", "sub_item", "response_option"]:
        df[c] = df[c].astype(str).replace("None", pd.NA)

    # Drop rows where response_option is missing
    df = df.dropna(subset=["response_option"])

    # Remove the 'No data to export' artifact rows
    df = df[~df["response_option"].str.contains("No data to export", na=False)]

    # Strip trailing whitespace from all string columns
    for c in df.select_dtypes(include="object").columns:
        df[c] = df[c].str.strip()

    # Replace curly quotes and other encoding artifacts
    for c in ["question_text", "sub_item", "response_option"]:
        df[c] = (df[c]
                 .str.replace("\u2019", "'", regex=False)
                 .str.replace("\u2018", "'", regex=False)
                 .str.replace("\u201c", '"', regex=False)
                 .str.replace("\u201d", '"', regex=False)
                 .str.replace("\u2026", "...", regex=False)
                 .str.replace("�", "'", regex=False))

    # Identify free-text questions (individual verbatim responses, mostly count=1)
    FREE_TEXT_QIDS = {"Q15", "Q26_3_TEXT", "Q27_7_TEXT", "Q5_6_TEXT"}
    is_freetext = df["question_id"].isin(FREE_TEXT_QIDS)
    freetext_df = df[is_freetext].copy()
    structured_df = df[~is_freetext].copy()

    return structured_df.reset_index(drop=True), freetext_df.reset_index(drop=True)


def melt_to_long(wide_df):
    """Melt the wide DataFrame into long format with segment_group, segment_value, count."""
    id_cols = ["question_id", "question_text", "sub_item", "response_option"]
    value_cols = [c for c in wide_df.columns if c.startswith(("moms_", "wsf_", "age_"))]

    long_df = wide_df.melt(
        id_vars=id_cols,
        value_vars=value_cols,
        var_name="segment_key",
        value_name="count",
    )

    # Parse segment_key into segment_group and segment_value
    def parse_segment(key):
        if key.startswith("moms_"):
            return "moms", key[5:]
        elif key.startswith("wsf_"):
            return "women_sports_fans", key[4:]
        elif key.startswith("age_"):
            return "age", key[4:]
        return "unknown", key

    long_df[["segment_group", "segment_value"]] = long_df["segment_key"].apply(
        lambda x: pd.Series(parse_segment(x))
    )
    long_df = long_df.drop(columns=["segment_key"])

    # Reorder columns
    long_df = long_df[["question_id", "question_text", "sub_item", "response_option",
                        "segment_group", "segment_value", "count"]]

    return long_df.reset_index(drop=True)


def main():
    print("Loading workbook...")
    wb = openpyxl.load_workbook(str(INPUT), data_only=True)
    ws = wb["1 q8"]

    print("Extracting raw rows...")
    raw_rows = load_raw_rows(ws)
    print(f"  Raw rows: {len(raw_rows)}")

    print("Parsing question blocks...")
    records = parse_all_questions(raw_rows)
    print(f"  Parsed records: {len(records)}")

    print("Building wide-format DataFrame...")
    wide_df, freetext_df = build_wide_df(records)
    print(f"  Structured wide shape: {wide_df.shape}")
    print(f"  Structured questions: {wide_df['question_id'].nunique()}")
    print(f"  Free-text rows: {len(freetext_df)} ({freetext_df['question_id'].nunique()} questions)")

    print("Melting to long format...")
    long_df = melt_to_long(wide_df)
    print(f"  Long shape: {long_df.shape}")

    # Save outputs
    wide_path = OUT_DIR / "survey_crosstabs_wide.csv"
    long_path = OUT_DIR / "survey_crosstabs_long.csv"
    freetext_path = OUT_DIR / "survey_freetext.csv"

    wide_df.to_csv(wide_path, index=False, encoding="utf-8-sig")
    long_df.to_csv(long_path, index=False, encoding="utf-8-sig")
    freetext_df.to_csv(freetext_path, index=False, encoding="utf-8-sig")

    print(f"\nSaved:")
    print(f"  {wide_path}")
    print(f"  {long_path}")
    print(f"  {freetext_path}")

    # Print summary stats
    print(f"\n--- Summary ---")
    print(f"Structured questions: {wide_df['question_id'].nunique()}")
    print(f"Wide rows (question x response combos): {len(wide_df)}")
    print(f"Long rows (wide x segments): {len(long_df)}")
    print(f"Free-text rows: {len(freetext_df)}")
    print(f"\nQuestion IDs:")
    for qid in wide_df["question_id"].unique():
        n = len(wide_df[wide_df["question_id"] == qid])
        print(f"  {qid}: {n} response options")
    print(f"\nFree-text Question IDs:")
    for qid in freetext_df["question_id"].unique():
        n = len(freetext_df[freetext_df["question_id"] == qid])
        print(f"  {qid}: {n} verbatim responses")

    # Quick data quality checks
    print(f"\n--- Quality Checks ---")
    print(f"Null counts in wide:")
    print(wide_df.isnull().sum().to_string())

    return wide_df, long_df, freetext_df


if __name__ == "__main__":
    wide_df, long_df, freetext_df = main()
